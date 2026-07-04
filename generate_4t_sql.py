import os
import bcrypt
import openpyxl
import time

folder_path = r"c:\Users\Rüstem\.gemini\antigravity\scratch\muro-mvz\4T Ana Gruplar ve Öğrencileri"
master_file = os.path.join(folder_path, "4t-Öğrenci Lİstesi.xlsx")
sql_output_file = r"c:\Users\Rüstem\.gemini\antigravity\scratch\muro-mvz\import_4t.sql"

def generate_custom_password(first_name, last_name, phone):
    # Türkçe karakterleri İngilizce karakterlere çevir
    tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    
    # İsim: Sadece ilk ad, küçük harf, ingilizce karakter
    f_name = first_name.split()[0].translate(tr_map).lower() if first_name else "ogrenci"
    
    # Telefonun son 2 hanesi
    p_str = "".join(filter(str.isdigit, str(phone)))
    last2 = p_str[-2:] if len(p_str) >= 2 else "00"
    
    # Soyismin ilk harfi, küçük harf, ingilizce karakter
    l_name = last_name.translate(tr_map).lower() if last_name else "x"
    l_char = l_name[0] if len(l_name) > 0 else "x"
    
    return f"{f_name}.{last2}.{l_char}"

sql_lines = []
sql_lines.append("BEGIN;")
sql_lines.append("-- 1. KISIM: OGRENCILERI (USERS) EKLE")

print("Ana öğrenci listesi okunuyor...")
wb = openpyxl.load_workbook(master_file, data_only=True)
sheet = wb.active

students = {}

for row in sheet.iter_rows(min_row=2, values_only=True):
    email = None
    name = ""
    phone = ""
    status = 1
    
    for idx, cell in enumerate(row):
        if cell and isinstance(cell, str) and "@" in cell:
            email = cell.strip()
            if idx >= 2:
                name = str(row[idx-2]).strip() if row[idx-2] else ""
                phone = str(row[idx-1]).strip() if row[idx-1] else ""
            if idx + 1 < len(row):
                status_val = row[idx+1]
                if status_val in [0, 1, 2, "0", "1", "2"]:
                    status = int(status_val)
            break
            
    if email:
        is_active = "true" if status == 1 else "false"
        name_parts = name.split(" ")
        last_name = name_parts[-1] if len(name_parts) > 1 else ""
        first_name = " ".join(name_parts[:-1]) if len(name_parts) > 1 else name
        
        if email not in students:
            # Şifreyi oluştur
            raw_pw = generate_custom_password(first_name, last_name, phone)
            # Hız için rounds=6 kullanıyoruz (Yine de güvenli)
            hashed_pw = bcrypt.hashpw(raw_pw.encode('utf-8'), bcrypt.gensalt(rounds=6)).decode('utf-8')
            
            students[email] = {
                "first_name": first_name.replace("'", "''"),
                "last_name": last_name.replace("'", "''"),
                "phone": phone.replace("'", "''"),
                "is_active": is_active,
                "hashed_pw": hashed_pw,
                "raw_pw_debug": raw_pw
            }

print(f"{len(students)} tekil öğrenci bulundu. SQL komutları hazırlanıyor...")

for email, data in students.items():
    safe_email = email.replace("'", "''")
    # Debug amaçlı üretilen şifreyi SQL dosyasında yorum satırı olarak da bırakalım
    sql_lines.append(f"-- Şifre Kuralı Test: {data['raw_pw_debug']} (Email: {safe_email})")
    
    sql = f"""
INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "IsActive", "CreatedAt", "IsDeleted", "FailedLoginCount")
SELECT gen_random_uuid(), '{data["first_name"]}', '{data["last_name"]}', '{safe_email}', '{safe_email}', '{data["phone"]}', '{data["hashed_pw"]}', 2, {data["is_active"]}, NOW(), false, 0
WHERE NOT EXISTS (SELECT 1 FROM "Users" WHERE "Email" = '{safe_email}');
"""
    sql_lines.append(sql.strip())

sql_lines.append("\n-- 2. KISIM: GRUPLARI EKLE")
print("Grup dosyaları taranıyor...")

groups = []
group_student_map = {} 

for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx") and filename != "4t-Öğrenci Lİstesi.xlsx":
        group_name = filename.replace(".xlsx", "").strip()
        groups.append(group_name)
        group_student_map[group_name] = []
        
        file_path = os.path.join(folder_path, filename)
        g_wb = openpyxl.load_workbook(file_path, data_only=True)
        g_sheet = g_wb.active
        for row in g_sheet.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and "@" in cell:
                    group_student_map[group_name].append(cell.strip())
                    break

for group in groups:
    safe_group = group.replace("'", "''")
    sql = f"""
INSERT INTO "Groups" ("Id", "Name", "IsDeleted", "CreatedAt")
SELECT gen_random_uuid(), '{safe_group}', false, NOW()
WHERE NOT EXISTS (SELECT 1 FROM "Groups" WHERE "Name" = '{safe_group}');
"""
    sql_lines.append(sql.strip())


sql_lines.append("\n-- 3. KISIM: ORENCILERI GRUPLARA BAGLA")
for group_name, emails in group_student_map.items():
    safe_group = group_name.replace("'", "''")
    for email in set(emails): 
        safe_email = email.replace("'", "''")
        sql = f"""
INSERT INTO "GroupMembers" ("Id", "GroupId", "UserId", "CreatedAt")
SELECT gen_random_uuid(), g."Id", u."Id", NOW()
FROM "Users" u, "Groups" g
WHERE u."Email" = '{safe_email}' AND g."Name" = '{safe_group}'
  AND NOT EXISTS (SELECT 1 FROM "GroupMembers" gm WHERE gm."GroupId" = g."Id" AND gm."UserId" = u."Id");
"""
        sql_lines.append(sql.strip())

sql_lines.append("COMMIT;")

with open(sql_output_file, "w", encoding="utf-8") as f:
    f.write("\n".join(sql_lines))

print(f"SQL dosyası oluşturuldu: {sql_output_file}")
