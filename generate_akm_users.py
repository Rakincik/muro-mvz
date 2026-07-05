import openpyxl
import uuid
import re

excel_file = "akademikmasa.xlsx"
sql_output_file = "import_akm_users.sql"

def clean_phone(p):
    if not p: return None
    s = str(p).strip().replace(" ", "")
    s = re.sub(r'[^0-9]', '', s)
    if s.startswith('90') and len(s) == 12: s = s[2:]
    if s.startswith('0'): s = s[1:]
    return s if len(s) == 10 else None

wb = openpyxl.load_workbook(excel_file, data_only=True)
s1 = wb["Sayfa1"]

users = {}
for row in s1.iter_rows(min_row=3, values_only=True):
    if not row[1]: continue
    name = str(row[1]).strip()
    phone = clean_phone(row[2])
    email = str(row[3]).strip().lower() if row[3] else None
    
    if not email or "@" not in email:
        safe_name = re.sub(r'[^a-z0-9]', '', name.lower().replace(" ", ""))
        email = f"{safe_name}@akademikmasa.com"

    parts = name.split()
    first_name = " ".join(parts[:-1]) if len(parts) > 1 else name
    last_name = parts[-1] if len(parts) > 1 else ""

    users[name] = {
        "id": str(uuid.uuid4()),
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "email": email
    }

s2 = wb["Sayfa2"]
group_assignments = []
for row in s2.iter_rows(min_row=3, values_only=True):
    if len(row) > 12 and row[11] and row[12]:
        grup_adi = str(row[11]).strip()
        ad_soyad = str(row[12]).strip()
        if grup_adi and ad_soyad != "Gösterilecek veri yok." and ad_soyad in users:
            group_assignments.append((users[ad_soyad]["id"], grup_adi))

with open(sql_output_file, "w", encoding="utf-8") as f:
    for u in users.values():
        pw = u["phone"] if u["phone"] else "123456"
        f.write(f"""
INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Phone", "PasswordHash", "Role", "IsActive", "CreatedAt", "Username")
SELECT '{u['id']}', '{u['first_name'].replace("'", "''")}', '{u['last_name'].replace("'", "''")}', '{u['email'].replace("'", "''")}', {f"'{u['phone']}'" if u['phone'] else 'NULL'}, '{pw}', 'Student', true, CURRENT_TIMESTAMP, '{u['email'].replace("'", "''")}'
FROM (SELECT 1) AS dummy
WHERE NOT EXISTS (SELECT 1 FROM "Users" WHERE "Email" = '{u['email'].replace("'", "''")}');
""")
        
    for user_id, group_name in group_assignments:
        f.write(f"""
INSERT INTO "GroupMembers" ("Id", "UserId", "GroupId", "Role", "Status", "AddedAt")
SELECT gen_random_uuid(), '{user_id}', "Id", 'Student', 'active', CURRENT_TIMESTAMP FROM "Groups" WHERE "Name" = '{group_name.replace("'", "''")}'
AND NOT EXISTS (
    SELECT 1 FROM "GroupMembers" 
    WHERE "UserId" = '{user_id}' 
    AND "GroupId" = (SELECT "Id" FROM "Groups" WHERE "Name" = '{group_name.replace("'", "''")}')
);
""")
print(f"Generated {len(users)} users and {len(group_assignments)} assignments in {sql_output_file}")
