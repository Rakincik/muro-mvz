import sys
import os
import argparse
import uuid
import datetime
import re
import psycopg2
import openpyxl

def clean_phone_number(phone):
    if not phone:
        return None
    phone_str = str(phone).strip()
    # Handle scientific notation or decimal parts from Excel (e.g. "5069993427.0")
    main_part = phone_str.split('.')[0].split(',')[0]
    # Keep only digits
    digits = "".join(c for c in main_part if c.isdigit())
    # Strip leading zeros
    while digits.startswith("0"):
        digits = digits[1:]
    # If 12 digits starting with 90, strip 90
    if len(digits) == 12 and digits.startswith("90"):
        digits = digits[2:]
    return digits if digits else None

def parse_name(full_name):
    if not full_name:
        return "", ""
    parts = str(full_name).strip().split()
    if len(parts) == 0:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def normalize(s):
    if not s:
        return ''
    s = str(s).lower().strip()
    s = s.replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def discover_connection_string():
    paths = [
        "src/MURO.API/appsettings.json",
        "appsettings.json",
        "../src/MURO.API/appsettings.json"
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    content = f.read()
                    match = re.search(r'"DefaultConnection"\s*:\s*"([^"]+)"', content)
                    if match:
                        conn_str = match.group(1)
                        params = {}
                        for item in conn_str.split(';'):
                            if '=' in item:
                                k, v = item.split('=', 1)
                                k_lower = k.strip().lower()
                                v_val = v.strip()
                                if k_lower == 'host': params['host'] = v_val
                                elif k_lower == 'port': params['port'] = int(v_val)
                                elif k_lower == 'database': params['dbname'] = v_val
                                elif k_lower == 'username': params['user'] = v_val
                                elif k_lower == 'password': params['password'] = v_val
                        if params:
                            return params, p
            except Exception:
                pass
    return None, None

def get_val(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]

def main():
    parser = argparse.ArgumentParser(description="MURO Excel Importer - Universal Version")
    parser.add_argument("--host", help="PostgreSQL host")
    parser.add_argument("--port", type=int, help="PostgreSQL port")
    parser.add_argument("--dbname", help="PostgreSQL database name")
    parser.add_argument("--user", help="PostgreSQL username")
    parser.add_argument("--password", help="PostgreSQL password")
    parser.add_argument("--file", default="mevzuatadam.xlsx", help="Excel file to import (default: mevzuatadam.xlsx)")
    parser.add_argument("--no-db", action="store_true", help="Skip db connection and only parse Excel data")
    parser.add_argument("--execute", action="store_true", help="Execute inserts (default is dry-run)")
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("            MURO LMS - DYNAMIC DATA IMPORTER")
    print("=" * 60)
    
    # Connection parameters discovery
    conn_params = {}
    disc_params, config_path = discover_connection_string()
    if disc_params:
        print(f"[*] Auto-discovered database config from: {config_path}")
        conn_params.update(disc_params)
        
    # Override with command line arguments if provided
    if args.host: conn_params['host'] = args.host
    if args.port: conn_params['port'] = args.port
    if args.dbname: conn_params['dbname'] = args.dbname
    if args.user: conn_params['user'] = args.user
    if args.password: conn_params['password'] = args.password
    
    # Default fallback if nothing discovered or provided
    if 'host' not in conn_params: conn_params['host'] = 'localhost'
    if 'port' not in conn_params: conn_params['port'] = 5432
    if 'dbname' not in conn_params: conn_params['dbname'] = 'muro_demo'
    if 'user' not in conn_params: conn_params['user'] = 'muro_user'
    if 'password' not in conn_params: conn_params['password'] = 'MuroDem0_2026!Str0ng'

    excel_file = args.file
    print(f"Reading Excel file '{excel_file}'...")
    if not os.path.exists(excel_file):
        print(f"[!] Error: '{excel_file}' not found in current directory.")
        sys.exit(1)
        
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        print(f"[!] Error opening Excel: {e}")
        sys.exit(1)
        
    if 'Sayfa1' not in wb.sheetnames:
        print("[!] Error: Excel must contain sheet: Sayfa1")
        sys.exit(1)
        
    s1 = wb['Sayfa1']
    rows1 = list(s1.iter_rows(values_only=True))
    
    # --- PARSE SAYFA1 HEADERS DYNAMICALLY ---
    header_row1 = 0
    for r_idx, row in enumerate(rows1[:10]):
        joined = " ".join([str(x).lower() for x in row if x is not None])
        if any(h in joined for h in ["ders", "onoff", "islem", "adi_soyadi", "telefon", "email"]):
            header_row1 = r_idx
            break
            
    headers1 = list(rows1[header_row1])
    
    def find_col(aliases):
        for idx, h in enumerate(headers1):
            if h and any(a == str(h).lower().strip() for a in aliases):
                return idx
        return None
        
    name_col = find_col(["ders", "adi_soyadi", "ad_soyad", "ad soyad", "adı soyadı"])
    phone_col = find_col(["onoff", "telefon", "tel", "cep"])
    email_col = find_col(["islem", "eposta", "email", "mail"])
    tc_col = find_col(["tc", "tckimlik", "tc_no", "tcno"])
    
    # Check for shifted column anomalies (e.g. name col actually has phones, phone col has emails)
    if name_col is not None and phone_col is not None:
        phone_in_name_col = 0
        email_in_phone_col = 0
        checks = 0
        for row in rows1[header_row1+1 : header_row1+15]:
            n_val = str(row[name_col]) if row[name_col] is not None else ""
            p_val = str(row[phone_col]) if row[phone_col] is not None else ""
            if n_val:
                checks += 1
                if n_val.isdigit() or len(re.sub(r'\D', '', n_val)) >= 7:
                    phone_in_name_col += 1
                if '@' in p_val:
                    email_in_phone_col += 1
                    
        if checks > 0 and phone_in_name_col / checks > 0.5:
            print("[*] Column shift detected: 'Name' column contains phone numbers, 'Phone' column contains emails.")
            # Map: Col 2 is phone, Col 3 is email, Name is missing in Sayfa1
            phone_col = name_col
            email_col = phone_col
            name_col = None

    # Parse Students in Sayfa1
    students_to_create = {}
    for row in rows1[header_row1+1:]:
        if not any(x is not None for x in row):
            continue
            
        raw_name = str(row[name_col]).strip() if name_col is not None and row[name_col] is not None else ""
        phone_val = str(row[phone_col]).strip() if phone_col is not None and row[phone_col] is not None else ""
        email_val = str(row[email_col]).strip() if email_col is not None and row[email_col] is not None else ""
        
        # Fallback if name has group name
        if not raw_name and email_val and "@" not in email_val:
            raw_name = email_val
            email_val = ""
            
        cleaned_phone = clean_phone_number(phone_val if phone_val else raw_name)
        if cleaned_phone:
            students_to_create[cleaned_phone] = {
                'name': raw_name,
                'phone': cleaned_phone,
                'email': email_val.lower() if email_val and "@" in email_val else f"student_{cleaned_phone}@muro.com"
            }

    # --- PARSE SAYFA2 HEADERS DYNAMICALLY ---
    has_sayfa2 = 'Sayfa2' in wb.sheetnames
    courses_to_create = set()
    course_modes = {}
    group_courses = []
    groups_to_create = set()
    memberships_to_create = []
    
    excel_groups = {} # maps Excel group_id to {name, parent_id, parent_name}

    if has_sayfa2:
        s2 = wb['Sayfa2']
        rows2 = list(s2.iter_rows(values_only=True))
        
        header_row2 = 0
        for r_idx, row in enumerate(rows2[:10]):
            joined = " ".join([str(x).lower() for x in row if x is not None])
            if "grup_id" in joined or "grup_adi" in joined or "ders" in joined:
                header_row2 = r_idx
                break
                
        headers2 = list(rows2[header_row2])
        
        # Split headers into Left Side (Courses) and Right Side (Students)
        grup_id_indices = [i for i, h in enumerate(headers2) if h and str(h).strip().lower() == 'grup_id']
        if len(grup_id_indices) >= 2:
            right_start_idx = grup_id_indices[1]
        else:
            right_start_idx = len(headers2) // 2
            for i in range(1, len(headers2)):
                if headers2[i] == 'grup_id' or headers2[i] == 'ad_soyad':
                    right_start_idx = i
                    break
                    
        left_headers = headers2[:right_start_idx]
        right_headers = headers2[right_start_idx:]
        
        def find_left_col(aliases):
            for idx, h in enumerate(left_headers):
                if h and any(a == str(h).lower().strip() for a in aliases):
                    return idx
            return None
            
        def find_right_col(aliases):
            for idx, h in enumerate(right_headers):
                if h and any(a == str(h).lower().strip() for a in aliases):
                    return right_start_idx + idx
            return None
            
        g_id_left_idx = find_left_col(['grup_id'])
        g_name_left_idx = find_left_col(['grup_adi'])
        g_parent_id_left_idx = find_left_col(['grup_ust_id', 'parent_id'])
        g_parent_name_left_idx = find_left_col(['grup_ust_adi', 'parent_name'])
        course_name_idx = find_left_col(['ders_adi', 'ders'])
        mode_idx = find_left_col(['mod'])
        
        g_id_right_idx = find_right_col(['grup_id'])
        g_name_right_idx = find_right_col(['grup_adi'])
        g_parent_id_right_idx = find_right_col(['grup_ust_id', 'parent_id'])
        g_parent_name_right_idx = find_right_col(['grup_ust_adi', 'parent_name'])
        student_name_idx = find_right_col(['ad_soyad', 'name', 'adi_soyadi', 'adı soyadı'])
        role_idx = find_right_col(['rol', 'role'])
        phone_idx_r = find_right_col(['telefon', 'phone'])
        email_idx_r = find_right_col(['email', 'eposta'])

        # Parse Left Table (Group-Course mapping)
        for r in rows2[header_row2+1:]:
            g_id = get_val(r, g_id_left_idx)
            g_name = get_val(r, g_name_left_idx)
            c_name = get_val(r, course_name_idx)
            mode = get_val(r, mode_idx)
            
            g_parent_id = get_val(r, g_parent_id_left_idx)
            g_parent_name = get_val(r, g_parent_name_left_idx)
            
            if g_id and g_name:
                g_id_str = str(g_id).strip()
                excel_groups[g_id_str] = {
                    'name': str(g_name).strip(),
                    'parent_id': str(g_parent_id).strip() if g_parent_id else None,
                    'parent_name': str(g_parent_name).strip() if g_parent_name else None
                }
                
                if c_name and str(c_name).strip() != "Gösterilecek veri yok.":
                    c_name_str = str(c_name).strip()
                    courses_to_create.add(c_name_str)
                    course_modes[c_name_str] = str(mode).strip() if mode else "Video"
                    group_courses.append((str(g_name).strip(), c_name_str, str(mode).strip() if mode else "Video", g_id_str))

        # Parse Right Table (Group-Student mapping)
        for r in rows2[header_row2+1:]:
            g_id_r = get_val(r, g_id_right_idx)
            g_name_r = get_val(r, g_name_right_idx)
            stud_name = get_val(r, student_name_idx)
            rol_str = get_val(r, role_idx)
            
            g_parent_id_r = get_val(r, g_parent_id_right_idx)
            g_parent_name_r = get_val(r, g_parent_name_right_idx)
            
            phone_r = get_val(r, phone_idx_r)
            email_r = get_val(r, email_idx_r)
            
            if g_id_r and g_name_r:
                g_id_r_str = str(g_id_r).strip()
                if g_id_r_str not in excel_groups:
                    excel_groups[g_id_r_str] = {
                        'name': str(g_name_r).strip(),
                        'parent_id': str(g_parent_id_r).strip() if g_parent_id_r else None,
                        'parent_name': str(g_parent_name_r).strip() if g_parent_name_r else None
                    }
                    
                if stud_name and str(stud_name).strip() != "Gösterilecek veri yok.":
                    memberships_to_create.append({
                        'group_excel_id': g_id_r_str,
                        'group_name': str(g_name_r).strip(),
                        'name': str(stud_name).strip(),
                        'role_str': str(rol_str).strip() if rol_str else 'Öğrenci',
                        'phone': str(phone_r).strip() if phone_r else '',
                        'email': str(email_r).strip() if email_r else ''
                    })

    groups_to_create = set(g['name'] for g in excel_groups.values())
    
    print(f"\n[+] Parsed Excel Data successfully:")
    print(f"    - Unique Groups  : {len(excel_groups)}")
    print(f"    - Unique Courses : {len(courses_to_create)}")
    print(f"    - Unique Students in User List: {len(students_to_create)}")
    print(f"    - Group-Course Mappings: {len(group_courses)}")
    print(f"    - Student-Group Memberships: {len(memberships_to_create)}")
    
    if args.no_db:
        print("\n[*] Running in local parse-only mode. Skipping DB connection.")
        sys.exit(0)
        
    print(f"\n[*] Connecting to database '{conn_params.get('dbname')}' on {conn_params.get('host')}:{conn_params.get('port')}...")
    try:
        conn = psycopg2.connect(**conn_params)
        cur = conn.cursor()
    except Exception as e:
        print(f"[!] Database Connection Failed: {e}")
        print("[!] Please check if your SSH tunnel is active or DB credentials are correct.")
        sys.exit(1)
        
    print("[+] Database connected successfully!")
    
    # Retrieve existing records to verify duplicates
    print("[*] Retrieving existing database records...")
    
    cur.execute('SELECT "Id", "Title" FROM "Courses" WHERE "IsDeleted" = False;')
    existing_courses = {r[1].lower().strip(): r[0] for r in cur.fetchall()}
    
    cur.execute('SELECT "Id", "Name" FROM "Groups" WHERE "IsDeleted" = False;')
    existing_groups = {r[1].lower().strip(): r[0] for r in cur.fetchall()}
    
    cur.execute('SELECT "Id", "Email", "Phone", "Username" FROM "Users";')
    users_db = cur.fetchall()
    existing_users_by_email = {r[1].lower().strip(): r[0] for r in users_db if r[1]}
    existing_users_by_phone = {r[2].strip(): r[0] for r in users_db if r[2]}
    existing_users_by_username = {r[3].lower().strip(): r[0] for r in users_db if r[3]}
    
    cur.execute('SELECT "CourseId", "GroupId" FROM "CourseGroups";')
    existing_course_groups = set((r[0], r[1]) for r in cur.fetchall())
    
    cur.execute('SELECT "UserId", "GroupId" FROM "GroupMembers";')
    existing_group_members = set((r[0], r[1]) for r in cur.fetchall())
    
    print("[+] Existing records loaded.")
    
    # Calculate write plan
    planned_students_new = sum(1 for p in students_to_create if p not in existing_users_by_phone and p not in existing_users_by_username)
    planned_courses_new = sum(1 for c in courses_to_create if c.lower().strip() not in existing_courses)
    
    planned_groups_new = 0
    for eg in excel_groups.values():
        if eg['name'].lower().strip() not in existing_groups:
            planned_groups_new += 1
            
    print("\n" + "=" * 50)
    print("                  PROPOSED WORKPLAN")
    print("=" * 50)
    print(f" 1. Create Students   : {planned_students_new} new to create, {len(students_to_create) - planned_students_new} already exist")
    print(f" 2. Create Courses    : {planned_courses_new} new to create, {len(courses_to_create) - planned_courses_new} already exist")
    print(f" 3. Create Groups     : {planned_groups_new} new to create (including hierarchy)")
    print(f" 4. Define Course-Group Mappings (CourseGroups)")
    print(f" 5. Define Student-Group Mappings (GroupMembers with correct Roles)")
    print("=" * 50)
    
    if not args.execute:
        print("\n[*] RUNNING IN DRY-RUN MODE. No changes will be written.")
        print(f"[*] To write changes to the database, run with: python import_data.py --file {excel_file} --execute")
        cur.close()
        conn.close()
        sys.exit(0)
        
    # Interactive confirmation prompt
    confirm = input("\n[?] Are you sure you want to execute this import on the database? (y/N): ")
    if confirm.lower().strip() not in ('y', 'yes'):
        print("[*] Operation cancelled.")
        cur.close()
        conn.close()
        sys.exit(0)
        
    print("\n[*] Starting transaction...")
    try:
        # STEP 1: CREATE STUDENTS (USERS)
        print("[1/5] Creating Students...")
        student_map = {} # cleaned phone -> UserId
        users_created = 0
        users_skipped = 0
        
        for phone, stud in students_to_create.items():
            email = stud['email']
            name = stud['name']
            first, last = parse_name(name)
            
            u_id = None
            if email.lower().strip() in existing_users_by_email:
                u_id = existing_users_by_email[email.lower().strip()]
            elif phone in existing_users_by_phone:
                u_id = existing_users_by_phone[phone]
            elif phone in existing_users_by_username:
                u_id = existing_users_by_username[phone]
                
            if u_id:
                student_map[phone] = u_id
                users_skipped += 1
            else:
                u_id = str(uuid.uuid4())
                student_map[phone] = u_id
                cur.execute(
                    'INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "StudentType", "IsActive", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                    (u_id, first, last, email, phone, phone, email, 'Student', 'Active', True, False, datetime.datetime.utcnow())
                )
                users_created += 1
        print(f"      -> Created: {users_created}, Skipped: {users_skipped}")
        
        # STEP 2: CREATE COURSES
        print("[2/5] Creating Courses...")
        course_map = {}
        courses_created = 0
        courses_skipped = 0
        for c_name in courses_to_create:
            c_key = c_name.lower().strip()
            if c_key in existing_courses:
                course_map[c_name] = existing_courses[c_key]
                courses_skipped += 1
            else:
                course_id = str(uuid.uuid4())
                course_map[c_name] = course_id
                cur.execute(
                    'INSERT INTO "Courses" ("Id", "Title", "IsDeleted", "IsPublished", "CourseType", "Mode", "Order", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                    (course_id, c_name, False, True, 'Online', 'Offline', 0, datetime.datetime.utcnow())
                )
                courses_created += 1
        print(f"      -> Created: {courses_created}, Skipped: {courses_skipped}")
        
        # STEP 3: CREATE GROUPS WITH HIERARCHY
        print("[3/5] Creating Groups with hierarchy...")
        group_map = {} # Excel group_id (str) -> DB Guid
        groups_created = 0
        groups_skipped = 0
        
        def ensure_group_created(e_id):
            nonlocal groups_created, groups_skipped
            if e_id in group_map:
                return group_map[e_id]
                
            g_info = excel_groups.get(e_id)
            if not g_info:
                return None
                
            g_name = g_info['name']
            parent_e_id = g_info['parent_id']
            
            parent_db_id = None
            if parent_e_id:
                parent_db_id = ensure_group_created(parent_e_id)
                
            g_key = g_name.lower().strip()
            if g_key in existing_groups:
                db_id = existing_groups[g_key]
                group_map[e_id] = db_id
                
                # Update ParentId if missing in database
                cur.execute(
                    'UPDATE "Groups" SET "ParentId" = %s WHERE "Id" = %s AND "ParentId" IS NULL',
                    (parent_db_id, db_id)
                )
                groups_skipped += 1
                return db_id
                
            db_id = str(uuid.uuid4())
            cur.execute(
                'INSERT INTO "Groups" ("Id", "Name", "ParentId", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s)',
                (db_id, g_name, parent_db_id, False, datetime.datetime.utcnow())
            )
            group_map[e_id] = db_id
            existing_groups[g_key] = db_id
            groups_created += 1
            return db_id

        for excel_id in excel_groups:
            ensure_group_created(excel_id)
            
        print(f"      -> Created: {groups_created}, Skipped/Updated: {groups_skipped}")
        
        # STEP 4: DEFINE COURSE-GROUP MAPPINGS (CourseGroups)
        print("[4/5] Defining Course-Group Mappings...")
        cg_created = 0
        cg_skipped = 0
        for g_name, c_name, mode, g_id_excel in group_courses:
            g_id = group_map.get(g_id_excel)
            c_id = course_map.get(c_name)
            if g_id and c_id:
                if (c_id, g_id) in existing_course_groups:
                    cg_skipped += 1
                else:
                    c_mode = 'Online' if mode == 'Canlı' else 'Offline'
                    cur.execute(
                        'INSERT INTO "CourseGroups" ("Id", "CourseId", "GroupId", "Mode", "AssignedAt") VALUES (%s, %s, %s, %s, %s)',
                        (str(uuid.uuid4()), c_id, g_id, c_mode, datetime.datetime.utcnow())
                    )
                    cg_created += 1
        print(f"      -> Created: {cg_created}, Skipped: {cg_skipped}")
        
        # STEP 5: DEFINE STUDENT-GROUP MAPPINGS (GroupMembers)
        print("[5/5] Defining Student-Group Mappings (with correct roles)...")
        
        # Map user names to IDs for loose lookup
        user_name_map = {}
        for phone, stud in students_to_create.items():
            u_id = student_map.get(phone)
            if u_id:
                user_name_map[normalize(stud['name'])] = u_id
                
        # Load existing users from DB into name map
        cur.execute('SELECT "Id", "FirstName", "LastName", "Email", "Phone" FROM "Users";')
        for r in cur.fetchall():
            db_uid, f_name, l_name, db_email, db_phone = r
            fullName = f"{f_name or ''} {l_name or ''}".strip()
            user_name_map[normalize(fullName)] = db_uid
            if db_email:
                user_name_map[normalize(db_email.split('@')[0])] = db_uid
            if db_phone:
                user_name_map[normalize(db_phone)] = db_uid
                
        members_created = 0
        members_skipped = 0
        
        for m in memberships_to_create:
            m_name = m['name']
            g_id = group_map.get(m['group_excel_id'])
            if not g_id:
                continue
                
            m_norm = normalize(m_name)
            u_id = user_name_map.get(m_norm)
            
            # Check for partial name match fallback
            if not u_id:
                for name_key, uid_val in user_name_map.items():
                    if name_key in m_norm or m_norm in name_key:
                        u_id = uid_val
                        break
                        
            # Auto-create if student is in group but completely missing from User List
            if not u_id:
                clean_name = m_name.replace('*', '').strip()
                username_slug = normalize(clean_name)
                if not username_slug:
                    username_slug = "student_" + str(uuid.uuid4())[:8]
                    
                phone_val = m['phone'] if m['phone'] else username_slug
                email_val = m['email'] if m['email'] else f"{username_slug}@muro.com"
                
                # Check if generated credentials already exist
                if email_val.lower().strip() in existing_users_by_email:
                    u_id = existing_users_by_email[email_val.lower().strip()]
                elif phone_val in existing_users_by_phone:
                    u_id = existing_users_by_phone[phone_val]
                else:
                    u_id = str(uuid.uuid4())
                    first, last = parse_name(clean_name)
                    cur.execute(
                        'INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "StudentType", "IsActive", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (u_id, first, last, email_val, phone_val, phone_val, email_val, 'Student', 'Active', True, False, datetime.datetime.utcnow())
                    )
                    user_name_map[m_norm] = u_id
                    existing_users_by_email[email_val.lower().strip()] = u_id
                    existing_users_by_phone[phone_val] = u_id
                    print(f"      [Auto-Created Student] {clean_name} (Username: {phone_val})")
            
            # Map role string correctly: Ynetici -> 0 (Admin), else -> 2 (Student)
            role_val = 2 # Student
            role_text = normalize(m['role_str'])
            if 'yonetici' in role_text or 'admin' in role_text:
                role_val = 0
                
            if u_id and g_id:
                if (u_id, g_id) in existing_group_members:
                    members_skipped += 1
                else:
                    cur.execute(
                        'INSERT INTO "GroupMembers" ("Id", "UserId", "GroupId", "Role", "Status", "AddedAt") VALUES (%s, %s, %s, %s, %s, %s)',
                        (str(uuid.uuid4()), u_id, g_id, role_val, 'active', datetime.datetime.utcnow())
                    )
                    members_created += 1
                    
        print(f"      -> Created: {members_created}, Skipped: {members_skipped}")
        
        # Commit transaction
        print("\n[*] Committing database transaction...")
        conn.commit()
        print("[+] SUCCESS: Data import completed successfully and committed to the database!")
        
    except Exception as db_err:
        print(f"\n[!] Database Transaction Error encountered: {db_err}")
        print("[!] Rolling back all changes in this transaction to preserve data integrity...")
        conn.rollback()
        print("[*] Rollback complete. No changes were saved to the database.")
        sys.exit(1)
        
    finally:
        cur.close()
        conn.close()

if __name__ == '__main__':
    main()
