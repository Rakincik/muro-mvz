import openpyxl
import random
import string
import re

excel_file = "atanıyorumhocam.xlsx"
sql_output_file = "import_trk_users.sql"

wb = openpyxl.load_workbook(excel_file, data_only=True)
s1 = wb["Sayfa1"]
s2 = wb["Sayfa2"]

# Parse Sayfa1: User details
users = []
seen_phones = set()

def clean_tr(text):
    tr_map = {'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c', 'İ': 'i', 'Ğ': 'g', 'Ü': 'u', 'Ş': 's', 'Ö': 'o', 'Ç': 'c'}
    for k, v in tr_map.items():
        text = text.replace(k, v)
    return text

for row in s1.iter_rows(min_row=2, values_only=True): 
    name = row[1]
    if not name or name == 'ders':
        continue
    
    phone = row[2]
    
    name_str = str(name).strip()
    name_key = name_str.lower()
    
    # Process Phone
    phone_str = ""
    if phone:
        phone_str = "".join(filter(str.isdigit, str(phone)))
        if phone_str.startswith("90") and len(phone_str) > 10:
            phone_str = phone_str[2:]
            
    # Guarantee unique phone string for username/email generation
    original_phone_str = phone_str
    if not phone_str:
        phone_str = f"random_{random.randint(100000, 999999)}"
        
    while phone_str in seen_phones:
        phone_str = f"{original_phone_str}_{random.randint(100, 999)}"
        
    seen_phones.add(phone_str)
    
    email_str = f"student_{phone_str}@muro.com"
            
    # Generate Password: isim.telefonunson2_hanesi.soyismin_ilk_harfi
    parts = name_str.split()
    if len(parts) >= 2:
        isim = "".join(parts[:-1]).lower()
        soyisim = parts[-1].lower()
    else:
        isim = name_str.lower()
        soyisim = name_str.lower() # Fallback
        
    isim = clean_tr(isim)
    soyisim_ilk_harf = clean_tr(soyisim[0]) if soyisim else "x"
    
    if original_phone_str and len(original_phone_str) >= 2:
        son2 = original_phone_str[-2:]
        password = f"{isim}.{son2}.{soyisim_ilk_harf}"
    else:
        random_pass = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
        password = f"{isim}.{random_pass}.{soyisim_ilk_harf}"
        
    users.append({
        'name_key': name_key,
        'name': name_str,
        'first_name': " ".join(parts[:-1]) if len(parts) >= 2 else name_str,
        'last_name': parts[-1] if len(parts) >= 2 else "-",
        'phone': phone_str, # unique guaranteed
        'actual_phone': original_phone_str,
        'email': email_str,
        'password': password
    })

# Parse Sayfa2: Group mappings
header_idx = -1
rows2 = list(s2.iter_rows(values_only=True))
for i, row in enumerate(rows2):
    if len(row) > 10 and row[10] == "ad_soyad":
        header_idx = i
        break

mappings = [] 
if header_idx != -1:
    for row in rows2[header_idx+1:]:
        if len(row) < 11:
            continue
            
        grup_adi = row[9]
        ad_soyad = row[10]
        
        if not grup_adi or not ad_soyad:
            continue
            
        name_key = str(ad_soyad).strip().lower()
        
        # Find user by name
        matched_user = next((u for u in users if u['name_key'] == name_key), None)
        if matched_user:
            mappings.append((matched_user, str(grup_adi).strip()))

# Generate SQL
sql_lines = ["BEGIN;"]
sql_lines.append("\n-- 1. KISIM: OGRENCILERI EKLE (Users)")

for u in users:
    safe_fn = u['first_name'].replace("'", "''")
    safe_ln = u['last_name'].replace("'", "''")
    safe_e = u['email'].replace("'", "''")
    safe_p = u['phone'].replace("'", "''")
    safe_pw = u['password'].replace("'", "''")
    
    sql = f"""
INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "PasswordHash", "Role", "IsActive", "CreatedAt", "Phone", "StudentType")
SELECT gen_random_uuid(), '{safe_fn}', '{safe_ln}', '{safe_e}', '{safe_p}', '{safe_pw}', 'Student', true, NOW(), '{safe_p}', 'Active'
WHERE NOT EXISTS (SELECT 1 FROM "Users" WHERE "Username" = '{safe_p}');
"""
    sql_lines.append(sql.strip())

sql_lines.append("\n-- 2. KISIM: OGRENCILERI GRUPLARA BAGLAMA (GroupMembers)")

for u, group in mappings:
    safe_p = u['phone'].replace("'", "''")
    safe_g = group.replace("'", "''")
    
    sql = f"""
INSERT INTO "GroupMembers" ("Id", "GroupId", "UserId", "Role", "Status", "AddedAt")
SELECT gen_random_uuid(), g."Id", u."Id", 2, 'active', NOW()
FROM "Users" u, "Groups" g
WHERE u."Username" = '{safe_p}' AND g."Name" = '{safe_g}'
  AND NOT EXISTS (SELECT 1 FROM "GroupMembers" gm WHERE gm."GroupId" = g."Id" AND gm."UserId" = u."Id");
"""
    sql_lines.append(sql.strip())

sql_lines.append("COMMIT;")

with open(sql_output_file, "w", encoding="utf-8") as f:
    f.write("\n".join(sql_lines))

print(f"Generated {sql_output_file} successfully. Processed {len(users)} unique users and {len(mappings)} user-group mappings.")
